import json
import pandas as pd
import tiktoken
from openai_api_calls import clinical_data_extractor
import time
from datetime import datetime
from openpyxl import load_workbook


def num_tokens_from_string(string: str, encoding_name: str) -> int:
    """Returns the number of tokens in a text string."""
    encoding = tiktoken.get_encoding(encoding_name)
    num_tokens = len(encoding.encode(string))
    return num_tokens


def get_estimated_cost(num_input_tokens, num_output_tokens, model):
    inp_cost = 0.0
    output_cost = 0.0
    if model == "gpt-4-1106-preview":
        inp_cost = 0.00001
        output_cost = 0.00003
    if model == "gpt-3.5-turbo-1106":
        inp_cost = 0.000001
        output_cost = 0.000002

    estimated_cost = inp_cost * num_input_tokens + output_cost * num_output_tokens

    return estimated_cost


def get_query_prompts(query, input_text):
    context = ""
    petition = ""
    dic = {}
    if query == "AP":
        labels = ["Diabetes",
                  "HTA",
                  "Fumador",
                  "Dislipemia",
                  "Hipotiroidismo",
                  "EPOC",
                  "Depresion",
                  "Renal",
                  "Fentanilo",
                  "Cardiopatia",
                  "Hipertiroidismo",
                  "Hepatopatia",
                  "Dependiente",
                  "Otras"]
        dic = {elem: "NO" for elem in labels}
        context = "Actúa como médico especialista en oncología radioterápica"
        petition = f"Utiliza el informe clínico proporcionado al final de este propmt para devolver en formato json el diccionario" \
                   + str(
            dic) + " con los valores \"SI\" o \"NO\" (para el campo \"Fumador\": \"SI\" si fuma, \"NO\" si nunca ha fumado, \"EX\" si es exfumador). " \
                   "Para el campo \"Otras\" devuelve una lista de comorbilidades encontradas que no puedan " \
                   "clasificarse en ninguna de las categorías de las claves del diccionario proporcionado, o vacío si no presenta otras comorbilidades." \
                   "Devuelve sólo el diccionario con los valores actualizados, NO AÑADIR NI MODIFICAR CLAVES. " \
                   f"Informe clínico: {input_text}"

    return context, petition


# Test que lee la información de antecedentes de fichero excel y llama a api de openia para obtener lista de comorbilidades
def exctract_data_from_reports(api_key, file_path, num_patients, model, query_type):

    start_time = time.time()
    context = ""
    petition = ""
    output_path = ""
    report_column = ""

    if query_type == "AP":
        report_column = "ANTECEDENTES PERSONALES"
        output_path = file_path.split('.xlsx')[-2] + '-' + model + '.xlsx'
        context, petition = get_query_prompts(query_type, "")

    # Read the specified sheet into a DataFrame
    df_all = pd.read_excel(file_path, sheet_name='DATOS')

    df = df_all.head(num_patients)

    json_data = []
    # Iterate through the rows of the DataFrame and print each row
    num_input_tokens = 0
    num_output_tokens = 0
    cont = 0
    for index, row in df.iterrows():
        try:
            json_output, system_content, user_content = clinical_data_extractor(api_key=api_key,
                                                                         model=model,
                                                                         context=context,
                                                                         petition=petition,
                                                                         input_text=row[report_column])

            # Compute the number of input and output tokens
            encoding_name = "cl100k_base"
            inp_tkns = num_tokens_from_string(system_content, encoding_name)
            inp_tkns += num_tokens_from_string(user_content, encoding_name)
            output_tkns = num_tokens_from_string(json.dumps(json_output), encoding_name)

            num_input_tokens += inp_tkns
            num_output_tokens += output_tkns

            # Add original text to output json
            json_output[report_column] = row[report_column]
            # Add patient ID
            patient_id = str(row['ID'])
            json_output['ID'] = patient_id
            json_data.append(json_output)
            cont += 1
            print(f"Informacion de paciente {patient_id} extraida correctamente ({cont})")
            print(json_output)
        except:
            print(f"Problema al extraer la información del paciente  {index} {row}")

    end_time = time.time()
    time_taken_minutes = (end_time - start_time) / 60

    # Compute cost from model and tokens
    estimated_cost = get_estimated_cost(num_input_tokens, num_output_tokens, model)

    # Print summary report
    print(f"Numero de textos analizados: {len(df)}")
    print(f"Numero de tokens consumidos: {num_input_tokens} input, {num_output_tokens} output")
    print(f"Coste estimado: {estimated_cost} $")
    print(f"Function completed in {time_taken_minutes:.2f} minutes.")
    df = pd.DataFrame(json_data)

    # Re-order to place ID n first column
    first_column = 'ID'
    if first_column in df.columns:
        reordered_columns = [first_column] + [col for col in df.columns if col != first_column]
        df = df[reordered_columns]

    # Replace "SÍ" with "SI" in the specified columns only
    data_columns = [col for col in df.columns if
                    ((col != 'ANTECEDENTES PERSONALES') or (col != 'ANTECEDENTES FAMILIARES'))]
    for column in data_columns:
        df[column] = df[column].replace("SÍ", "SI")

    save_results_to_excel(df, output_path, model, context, petition, time_taken_minutes, estimated_cost)

def save_results_to_excel(df, output_path, model, context, petition, time, cost):
    # Save results to "DATOS" excel sheet
    df.to_excel(output_path, index=False, sheet_name=f"DATOS")

    # Write summary report to "InfoConsulta" sheet
    workbook = load_workbook(filename=output_path)
    # Check if the 'prompts' sheet already exists; if not, create it
    if "InfoConsulta" not in workbook.sheetnames:
        workbook.create_sheet("InfoConsulta")
    sheet = workbook["InfoConsulta"]
    sheet['A1'] = "Modelo:"
    sheet['B1'] = model
    sheet['A2'] = 'Contexto:'
    sheet['B2'] = str(context)
    sheet['A3'] = 'Peticion:'
    sheet['B3'] = str(petition)
    sheet['A4'] = "Numero pacientes::"
    sheet['B4'] = len(df)
    sheet['A5'] = "Tiempo de ejecución (min):"
    sheet['B5'] = time
    sheet['A6'] = "Coste estimado ($):"
    sheet['B6'] = cost
    now = datetime.now()
    formatted_datetime = now.strftime('%Y-%m-%d %H:%M:%S')
    sheet['A7'] = "Fecha y hora consulta:"
    sheet['B7'] = formatted_datetime
    workbook.save(filename=output_path)

def main():
    api_file = "openai_key.txt"
    with open(api_file, 'r') as file:
        api_key = file.read().strip()


    file_path='./data/Dataset_AP.xlsx'
    # Numero de pacientes a analizar
    num_pacientes = 5
    model = "gpt-3.5-turbo-1106"         # "gpt-3.5-turbo-1106" , "gpt-4-1106-preview"
    query_type = "AP"     # Comorbidities and lifestyle risk factor from personal history report
    exctract_data_from_reports(api_key, file_path, num_pacientes, model, query_type)

if __name__ == '__main__':
    main()

